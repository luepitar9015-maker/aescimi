import openpyxl
from openpyxl.styles import PatternFill, Font
import pandas as pd
import os

def main():
    excel_path = r"C:\Users\Usuario\Documents\UNIDAD DE CORRESPONDENCIA PENDIENTE\PENDIENTES PRODUCIDAS_EDITADO.xlsx"
    report_path = r"C:\Users\Usuario\Documents\UNIDAD DE CORRESPONDENCIA PENDIENTE\COMUNICACIONES PRODUCIDAS\REPORTE_CONSULTA_ONBASE.xlsx"
    
    print(f"Cargando reporte de consulta: {report_path}")
    df_report = pd.read_excel(report_path)
    
    # Crear diccionario de resultados
    status_map = {}
    for _, row in df_report.iterrows():
        status_map[str(row['Radicado']).strip()] = row['Estado_OnBase']
        
    print(f"Cargando libro original de Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Encontrar las cabeceras de columna
    col_radicado_idx = None
    col_estado_idx = None
    
    for col_idx in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=1, column=col_idx).value
        if cell_val == "COMUNICACIONES PRODUCIDAS":
            col_radicado_idx = col_idx
        elif cell_val == "ESTADO":
            col_estado_idx = col_idx
            
    if not col_radicado_idx or not col_estado_idx:
        print("ERROR: No se encontraron las columnas 'COMUNICACIONES PRODUCIDAS' o 'ESTADO' en la primera fila.")
        return
        
    # Estilos estándar de Excel (Conditional Formatting)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(color="006100", bold=True)
    
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006", bold=True)
    
    updated_count = 0
    
    # Iterar sobre las filas a partir de la fila 2
    for row_idx in range(2, ws.max_row + 1):
        rad_val = str(ws.cell(row=row_idx, column=col_radicado_idx).value or "").strip()
        
        if rad_val in status_map:
            estado_cell = ws.cell(row=row_idx, column=col_estado_idx)
            status = status_map[rad_val]
            
            if status == "APARECE":
                estado_cell.value = "ya radicadas"
                estado_cell.fill = green_fill
                estado_cell.font = green_font
            else:
                estado_cell.value = "hacer acta"
                estado_cell.fill = red_fill
                estado_cell.font = red_font
                
            updated_count += 1
            
    print(f"Filas actualizadas y coloreadas: {updated_count}")
    
    # Guardar el archivo Excel
    print(f"Guardando cambios en: {excel_path}")
    wb.save(excel_path)
    print("¡Listo! El archivo ha sido actualizado exitosamente.")

if __name__ == '__main__':
    main()
