import openpyxl
import glob
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

def search_files():
    excel_files = glob.glob(r"C:\Users\Usuario\Downloads\*.xlsx")
    print(f"Buscando en {len(excel_files)} archivos Excel en Downloads...")

    for fpath in excel_files:
        if "~$" in fpath:
            continue
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            for sname in wb.sheetnames:
                ws = wb[sname]
                found = 0
                for row in ws.iter_rows(values_only=True):
                    row_str = " ".join([str(c) for c in row if c is not None]).lower()
                    if "viviana" in row_str or "villamizar" in row_str:
                        found += 1
                if found > 0:
                    print(f"-> ENCONTRADO en archivo: {fpath} | Hoja: {sname} | Coincidencias: {found}")
        except Exception as e:
            pass

if __name__ == "__main__":
    search_files()
